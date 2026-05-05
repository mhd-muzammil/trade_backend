import os
from fastapi import FastAPI, File, UploadFile, HTTPException, Response,Request
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
import numpy as np
import io
import json
from datetime import datetime, timedelta, timezone
from typing import List
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from sqlalchemy import create_engine, Column, Integer, String, Text, DateTime
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker
from dotenv import load_dotenv
from fastapi import Request

IST = timezone(timedelta(hours=5, minutes=30))

def get_ist_now():
    return datetime.now(IST).replace(tzinfo=None)

# Load environment variables
load_dotenv()

DATABASE_URL = os.getenv("DATABASE_URL")

# Normalize postgres URL format
if DATABASE_URL and DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

from sqlalchemy.orm import declarative_base

Base = declarative_base()

class TradeRecord(Base):
    __tablename__ = "trade_records"
    id = Column(Integer, primary_key=True, index=True)
    ticket_no = Column(String(50))
    case_id = Column(String(50))
    status = Column(String(50))
    product_name = Column(Text)
    asp_city = Column(String(100))
    wip_aging = Column(Integer)
    created_at = Column(DateTime, default=get_ist_now)

class FileHistory(Base):
    __tablename__ = "file_history"
    id = Column(Integer, primary_key=True, index=True)
    filename = Column(String(255))
    action = Column(String(50))
    total_records = Column(Integer)
    filtered_records = Column(Integer)
    created_at = Column(DateTime, default=get_ist_now)

try:
    if not DATABASE_URL:
        raise ValueError("DATABASE_URL is not set")
    engine = create_engine(DATABASE_URL)
    # Test connection
    with engine.connect() as conn:
        pass
    SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    Base.metadata.create_all(bind=engine)
except Exception as e:
    print(f"PostgreSQL connection failed ({e}). Falling back to SQLite.")
    DATABASE_URL = "sqlite:///./trade.db"
    engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
    SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    Base.metadata.create_all(bind=engine)

app = FastAPI(title="Trade Report API")

# CORS — configurable via CORS_ORIGINS env var (comma-separated), defaults to allow all
cors_origins = [
    origin.strip()
    for origin in os.getenv("CORS_ORIGINS", "https://trade-backend-qk0h.onrender.com/,http://localhost:5173").split(",")
    if origin.strip()
]
allow_all_origins = cors_origins == ["*"]
app.add_middleware(
    CORSMiddleware,
    allow_origins=cors_origins,
    allow_credentials=not allow_all_origins,
    allow_methods=["*"],
    allow_headers=["*"],
    # ADD Content-Disposition HERE:
    expose_headers=["X-Records-Processed", "X-Records-Filtered", "X-File-Stats", "X-City-Stats", "Content-Disposition"],
)

@app.get("/health")
@app.head("/health")
async def health_check():
    return {"status": "healthy"}

TARGET_COLUMNS = [
    "Ticket No",
    "Case Id",
    "Current Remarks",
    "WIP Aging",
    "WIP Aging Category",
    "Status",
    "HP Owner",
    "Product Name",
    "Product Serial No",
    "Product Type",
    "ASP City"
]


def process_single_file(contents: bytes, filename: str) -> dict:
    """Process a single Flex WIP Excel file and return filtered DataFrame + stats."""
    excel_file = pd.ExcelFile(io.BytesIO(contents))

    # Read from 'Data' sheet if it exists, else the first sheet
    sheet_name = 'Data' if 'Data' in excel_file.sheet_names else excel_file.sheet_names[0]
    df = excel_file.parse(sheet_name)

    total_records = len(df)

    # Filter for "01-Trade" in "WO OTC Code"
    if "WO OTC Code" not in df.columns:
        raise ValueError(f"'WO OTC Code' column not found in '{filename}'.")

    df_filtered = df[df["WO OTC Code"] == "01-Trade"].copy()
    filtered_records = len(df_filtered)

    # Ensure all 11 target columns exist, fill with blanks if missing
    for col in TARGET_COLUMNS:
        if col not in df_filtered.columns:
            df_filtered[col] = ""

    # Extract the specific 11 fields
    df_final = df_filtered[TARGET_COLUMNS]

    # Calculate city breakdown
    city_counts = {}
    if "ASP City" in df_final.columns:
        counts = df_final["ASP City"].fillna("Unknown").value_counts().to_dict()
        city_counts = {str(k): int(v) for k, v in counts.items()}

    return {
        "df": df_final,
        "total_records": total_records,
        "filtered_records": filtered_records,
        "filename": filename,
        "city_counts": city_counts
    }


def add_pivot_table_sheet(writer, df_source):
    """Add a 'Pivot Table' sheet with a working ASP City dropdown filter.

    Layout mirrors a real Excel PivotTable:
      - Rows    = Status
      - Columns = WIP Aging (sorted numerically)
      - Values  = Count of Ticket No
      - Filter  = ASP City (dropdown in B1)

    A hidden 'PivotData' sheet holds the raw data; every value cell is a
    COUNTIFS formula that reacts to the ASP City selection in B1.
    """
    from openpyxl.worksheet.datavalidation import DataValidation

    required_cols = ['Status', 'WIP Aging', 'Ticket No', 'ASP City']
    if not all(col in df_source.columns for col in required_cols):
        return

    # ---- Prepare source data ----
    df_pv = df_source[required_cols].copy()
    df_pv = df_pv.dropna(subset=['Status', 'Ticket No'])
    df_pv = df_pv[df_pv['Status'].astype(str).str.strip() != '']
    if df_pv.empty:
        return

    df_pv['WIP Aging'] = pd.to_numeric(df_pv['WIP Aging'], errors='coerce')
    df_pv = df_pv.dropna(subset=['WIP Aging'])
    df_pv['WIP Aging'] = df_pv['WIP Aging'].astype(int)

    df_pv['ASP City'] = df_pv['ASP City'].fillna('Unknown').astype(str).str.strip()
    df_pv.loc[df_pv['ASP City'] == '', 'ASP City'] = 'Unknown'

    unique_statuses = sorted(df_pv['Status'].unique())
    unique_wip = sorted(df_pv['WIP Aging'].unique())
    unique_cities = sorted(df_pv['ASP City'].unique())

    # ---- Write raw data to hidden helper sheet ----
    # Column order: A=Status, B=WIP Aging, C=Ticket No, D=ASP City
    df_pv = df_pv[['Status', 'WIP Aging', 'Ticket No', 'ASP City']]
    df_pv.to_excel(writer, index=False, sheet_name='PivotData')
    ws_data = writer.sheets['PivotData']
    ws_data.sheet_state = 'hidden'

    n = len(df_pv) + 1  # last data row (row 1 = header)

    # Write city list in column F for the Data Validation dropdown
    ws_data.cell(row=1, column=6, value='CityFilter')
    ws_data.cell(row=2, column=6, value='(All)')
    for i, city in enumerate(unique_cities):
        ws_data.cell(row=3 + i, column=6, value=city)
    last_city_row = 2 + len(unique_cities)

    # Named ranges for COUNTIFS formulas
    sr = f"PivotData!$A$2:$A${n}"   # Status
    wr = f"PivotData!$B$2:$B${n}"   # WIP Aging
    cr = f"PivotData!$D$2:$D${n}"   # ASP City

    # ---- Create Pivot Table worksheet at the front ----
    ws = writer.book.create_sheet('Pivot Table', 0)
    writer.book.active = ws

    # Styling tokens
    hdr_font = Font(bold=True, size=11)
    hdr_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    gt_fill  = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    ctr = Alignment(horizontal='center')

    # --- Row 1: ASP City dropdown filter ---
    ws['A1'] = 'ASP City'
    ws['A1'].font = Font(bold=True, color='006100')
    ws['B1'] = '(All)'
    ws['B1'].font = Font(italic=True)

    dv = DataValidation(
        type='list',
        formula1=f'PivotData!$F$2:$F${last_city_row}',
        allow_blank=False,
    )
    dv.prompt = 'Select an ASP City or (All)'
    dv.promptTitle = 'ASP City Filter'
    ws.add_data_validation(dv)
    dv.add(ws['B1'])

    # --- Row 3: titles ---
    ws['A3'] = 'Count of Ticket No'
    ws['A3'].font = hdr_font
    ws['B3'] = 'Column Labels'
    ws['B3'].font = hdr_font

    # --- Row 4: column headers ---
    num_wip = len(unique_wip)
    gt_col = num_wip + 2  # Grand Total column index (1-based)

    ws.cell(row=4, column=1, value='Row Labels')
    for i, wip_val in enumerate(unique_wip):
        ws.cell(row=4, column=i + 2, value=wip_val)
    ws.cell(row=4, column=gt_col, value='Grand Total')

    for c in range(1, gt_col + 1):
        cell = ws.cell(row=4, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = thin
        cell.alignment = ctr

    # --- Helper to build a COUNTIFS formula for any cell ---
    def _formula(r, c, is_gt_col=False, is_gt_row=False):
        cl = get_column_letter(c)
        if is_gt_row and is_gt_col:
            all_f  = f'COUNTA({sr})'
            city_f = f'COUNTIF({cr},$B$1)'
        elif is_gt_row:
            all_f  = f'COUNTIF({wr},{cl}$4)'
            city_f = f'COUNTIFS({wr},{cl}$4,{cr},$B$1)'
        elif is_gt_col:
            all_f  = f'COUNTIF({sr},$A{r})'
            city_f = f'COUNTIFS({sr},$A{r},{cr},$B$1)'
        else:
            all_f  = f'COUNTIFS({sr},$A{r},{wr},{cl}$4)'
            city_f = f'COUNTIFS({sr},$A{r},{wr},{cl}$4,{cr},$B$1)'
        inner = f'IF($B$1="(All)",{all_f},{city_f})'
        return f'=IF({inner}=0,"",{inner})'

    # --- Data rows (row 5 … 5+N-1) ---
    for row_off, status in enumerate(unique_statuses):
        r = 5 + row_off
        ws.cell(row=r, column=1, value=status).border = thin
        for col_off in range(num_wip):
            c = 2 + col_off
            cell = ws.cell(row=r, column=c)
            cell.value = _formula(r, c)
            cell.border = thin
            cell.alignment = ctr
        # Grand Total column
        cell = ws.cell(row=r, column=gt_col)
        cell.value = _formula(r, gt_col, is_gt_col=True)
        cell.border = thin
        cell.alignment = ctr

    # --- Grand Total row ---
    gt_row = 5 + len(unique_statuses)
    ws.cell(row=gt_row, column=1, value='Grand Total').border = thin
    for col_off in range(num_wip):
        c = 2 + col_off
        cell = ws.cell(row=gt_row, column=c)
        cell.value = _formula(gt_row, c, is_gt_row=True)
        cell.border = thin
        cell.alignment = ctr
    # Corner cell (Grand Total × Grand Total)
    cell = ws.cell(row=gt_row, column=gt_col)
    cell.value = _formula(gt_row, gt_col, is_gt_col=True, is_gt_row=True)
    cell.border = thin
    cell.alignment = ctr

    # Grand Total row styling
    for c in range(1, gt_col + 1):
        cell = ws.cell(row=gt_row, column=c)
        cell.font = hdr_font
        cell.fill = gt_fill

    # --- Column widths ---
    ws.column_dimensions['A'].width = 32
    for c in range(2, gt_col):
        ws.column_dimensions[get_column_letter(c)].width = 6
    ws.column_dimensions[get_column_letter(gt_col)].width = 13


@app.post("/api/process-report")
async def process_report(file: UploadFile = File(...)):
    """Process a single Flex WIP report."""
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")

    try:
        contents = await file.read()
        result = process_single_file(contents, file.filename)
        df_final = result["df"]

        # Generate the new Excel file
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_final.to_excel(writer, index=False, sheet_name='Trade Report')
            ws_trade = writer.sheets['Trade Report']
            max_col = ws_trade.max_column
            max_row = ws_trade.max_row
            col_letter = get_column_letter(max_col)
            ws_trade.auto_filter.ref = f"A1:{col_letter}{max_row}"
            add_pivot_table_sheet(writer, df_final)

        output.seek(0)

        date_str = datetime.now().strftime("%d-%m-%Y")
        filename = f"Trade_Report_{date_str}.xlsx"

        city_counts_json = json.dumps(result.get("city_counts", {}))

        # Save to file history
        db = SessionLocal()
        try:
            FileHistory.__table__.create(bind=engine, checkfirst=True)
            import_record = FileHistory(
                filename=file.filename,
                action="Import",
                total_records=result["total_records"],
                filtered_records=result["filtered_records"],
                created_at=get_ist_now()
            )
            export_record = FileHistory(
                filename=filename,
                action="Export",
                total_records=result["total_records"],
                filtered_records=result["filtered_records"],
                created_at=get_ist_now()
            )
            db.add(import_record)
            db.add(export_record)
            db.commit()
        except Exception as db_err:
            print(f"Error logging file history: {db_err}")
            db.rollback()
        finally:
            db.close()

        headers = {
            'Content-Disposition': f'attachment; filename="{filename}"',
            'X-Records-Processed': str(result["total_records"]),
            'X-Records-Filtered': str(result["filtered_records"]),
            'X-City-Stats': city_counts_json
        }

        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers
        )

    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/process-multiple")
async def process_multiple_reports(files: List[UploadFile] = File(...)):
    """Process multiple Flex WIP reports and combine into one Trade Report."""
    if not files:
        raise HTTPException(status_code=400, detail="No files uploaded.")

    all_dfs = []
    file_stats = []
    total_processed = 0
    total_filtered = 0
    total_city_counts = {}

    for file in files:
        if not file.filename.endswith('.xlsx'):
            raise HTTPException(
                status_code=400,
                detail=f"'{file.filename}' is not a .xlsx file. Only .xlsx files are supported."
            )

        try:
            contents = await file.read()
            result = process_single_file(contents, file.filename)
            all_dfs.append(result["df"])
            total_processed += result["total_records"]
            total_filtered += result["filtered_records"]
            
            for k, v in result.get("city_counts", {}).items():
                total_city_counts[k] = total_city_counts.get(k, 0) + v

            file_stats.append({
                "filename": result["filename"],
                "total": result["total_records"],
                "filtered": result["filtered_records"]
            })
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Error processing '{file.filename}': {str(e)}")

    # Combine all DataFrames
    combined_df = pd.concat(all_dfs, ignore_index=True)

    # Generate the combined Excel file
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        combined_df.to_excel(writer, index=False, sheet_name='Trade Report')
        ws_trade = writer.sheets['Trade Report']
        max_col = ws_trade.max_column
        max_row = ws_trade.max_row
        col_letter = get_column_letter(max_col)
        ws_trade.auto_filter.ref = f"A1:{col_letter}{max_row}"
        add_pivot_table_sheet(writer, combined_df)

    output.seek(0)

    date_str = datetime.now().strftime("%d-%m-%Y")
    filename = f"Trade_Report_Combined_{date_str}.xlsx"

    city_counts_json = json.dumps(total_city_counts)

    # Save to file history
    db = SessionLocal()
    try:
        FileHistory.__table__.create(bind=engine, checkfirst=True)
        for file in files:
            matched_stat = next((s for s in file_stats if s["filename"] == file.filename), None)
            tot = matched_stat["total"] if matched_stat else 0
            filt = matched_stat["filtered"] if matched_stat else 0
            import_record = FileHistory(
                filename=file.filename,
                action="Import",
                total_records=tot,
                filtered_records=filt,
                created_at=get_ist_now()
            )
            db.add(import_record)

        export_record = FileHistory(
            filename=filename,
            action="Export",
            total_records=total_processed,
            filtered_records=total_filtered,
            created_at=get_ist_now()
        )
        db.add(export_record)
        db.commit()
    except Exception as db_err:
        print(f"Error logging file history in multiple: {db_err}")
        db.rollback()
    finally:
        db.close()

    headers = {
        'Content-Disposition': f'attachment; filename="{filename}"',
        'X-Records-Processed': str(total_processed),
        'X-Records-Filtered': str(total_filtered),
        'X-File-Stats': json.dumps(file_stats),
        'X-City-Stats': city_counts_json
    }

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )


@app.api_route("/api/history", methods=["GET", "HEAD"])
@app.api_route("/api/history/", methods=["GET", "HEAD"])
async def get_history(request: Request):
    db = SessionLocal()
    try:
        FileHistory.__table__.create(bind=engine, checkfirst=True)
        records = db.query(FileHistory).order_by(FileHistory.created_at.desc()).all()
        return [
            {
                "id": r.id,
                "filename": r.filename,
                "action": r.action,
                "total_records": r.total_records,
                "filtered_records": r.filtered_records,
                "created_at": r.created_at.isoformat() if r.created_at else None
            }
            for r in records
        ]
    except Exception as e:
        print(f"Error fetching history: {e}")
        return []
    finally:
        db.close()

@app.get("/api/check-history")
async def check_history_endpoint():
    return {"status": "working", "endpoint": "/api/history is active"}


@app.delete("/api/history")
async def clear_history():
    db = SessionLocal()
    try:
        FileHistory.__table__.create(bind=engine, checkfirst=True)
        count = db.query(FileHistory).count()
        db.query(FileHistory).delete()
        db.commit()
        return {"message": f"Cleared {count} history records"}
    except Exception as e:
        db.rollback()
        print(f"Error clearing history: {e}")
        return {"message": "Failed to clear history"}
    finally:
        db.close()


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8001, reload=True)
